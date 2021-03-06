#!/usr/bin/env python2

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.cell import get_column_letter as gcl, Cell
from openpyxl.style import Alignment, NumberFormat, Fill, Color, Font, Border

#### WARNING: OpenPYXl has a bug that breaks office 2007 compatibility:
####   https://bitbucket.org/ericgazoni/openpyxl/issue/109/simple-formula-write-not-working
#### OpenPYXl 1.6.1 still has this bug, so I've had to patch library manually

class ReportFormatter:
    COLUMN_LOCATION  = 1
    COLUMN_ORDERNO   = 2
    COLUMN_DATE      = 3
    COLUMN_ITEM      = 4
    COLUMN_PARTNO    = 5
    COLUMN_QUANTITY  = 6
    COLUMN_UNITPRICE = 7
    COLUMN_EXTENDED  = 8
    COLUMN_SHIPPING  = 9
    COLUMN_HST       = 10
    COLUMN_TOTAL     = 11

    HEADER = ['Location', 'Order No', 'Date', 'Item', 'Part Number', 'Quantity', 'Unit Price', 'Extended', 'Shipping', 'HST', 'Total']
    COLUMN_WIDTH = [35, 17, 11, 34, 15, 9, 10]
    TOTALS = ['Extended', 'Freight', 'HST', 'Total']

    BORDERS = ['left', 'right', 'top', 'bottom']

    def __init__(self, title = None):
        self.wb = Workbook()
        self.ws = self.wb.worksheets[0]
        self.row = 1

        if title is not None:
            self.ws.header_footer.center_header.text = '&"-,Bold"%s' % (title.replace('&', '&&'),)

    def cell(self, x, y, value = None, bold = None, halign = None, formula = None, format = None, color = None, border = None, wrap_text = None):
        cell = self.ws.cell('%s%s' % (gcl(x), y))

        if format is not None:
            cell.style.number_format.format_code = format

        if value is not None:
            if formula:
                cell.set_value_explicit(value, Cell.TYPE_FORMULA)
            else:
                cell.value = value

        if halign is not None:
            cell.style.alignment.horizontal = halign

        if color is not None:
            cell.style.font.color.index = color

        if border is not None:
            for side in self.BORDERS:
                getattr(cell.style.borders, side).border_style = border

        if wrap_text is not None:
            cell.style.alignment.wrap_text = wrap_text

        if bold is not None:
            cell.style.font.bold = bold

        return cell

    def _add_header(self, offset = None, headers = None):
        if offset is None:
            offset = 0

        if headers is None:
            headers = self.HEADER

        for idx, i in enumerate(headers):
            self.cell(offset + idx + 1, self.row, value=i, bold=True)

    def section(self, location):
        self._add_header()
        self.row += 1
        self.cell(self.COLUMN_LOCATION, self.row, value=location, bold=True)

        self._remember_section(location)

    def _remember_section(self, location):
        self.location_name = location
        self.section_row = self.row

    def order(self, order):
        self.cell(self.COLUMN_ORDERNO, self.row, value=order.name)
        self.cell(self.COLUMN_DATE, self.row, value=order.due_date, format=NumberFormat.FORMAT_DATE_XLSX14)

    def items(self, items):
        for item in items:
            self.item(item)

    def item(self, item):
        self.cell(self.COLUMN_ITEM, self.row, wrap_text=True, value=str(item))
        self.cell(self.COLUMN_PARTNO, self.row, value=item.inventory_history.product.part_number)

        self.cell(self.COLUMN_QUANTITY, self.row, value=item.inventory_history.amount, halign=Alignment.HORIZONTAL_RIGHT)

        self.cell(self.COLUMN_UNITPRICE, self.row, format=NumberFormat.FORMAT_CURRENCY_USD_SIMPLE, value=item.inventory_history.product.price)

        self.cell(
            self.COLUMN_EXTENDED, self.row,
            '=%s%d * %s%d' % (
                gcl(self.COLUMN_UNITPRICE), self.row,
                gcl(self.COLUMN_QUANTITY), self.row,
            ),
            formula=True,
            format=NumberFormat.FORMAT_CURRENCY_USD_SIMPLE
        )

        self.cell(self.COLUMN_SHIPPING, self.row, format=NumberFormat.FORMAT_CURRENCY_USD_SIMPLE)

        self.cell(
            self.COLUMN_HST, self.row,
            '=SUM(%s%d, %s%d) * 0.13' % (
                gcl(self.COLUMN_EXTENDED), self.row,
                gcl(self.COLUMN_SHIPPING), self.row,
            ),
            format=NumberFormat.FORMAT_CURRENCY_USD_SIMPLE,
            formula=True
        )

        self.cell(
            self.COLUMN_TOTAL, self.row,
            '=SUM(%s%d, %s%d, %s%d)' % (
                gcl(self.COLUMN_EXTENDED), self.row,
                gcl(self.COLUMN_SHIPPING), self.row,
                gcl(self.COLUMN_HST), self.row
            ),
            format=NumberFormat.FORMAT_CURRENCY_USD_SIMPLE,
            formula=True
        )

        self.row += 1

    def _row_color(self, color = '606060', row = None):
        if row is None:
            row = self.row

        for i in xrange(1, len(self.HEADER) + 1):
            cell = self.cell(i, row)

            if color is None:
                cell.style.fill.fill_type = Fill.FILL_NONE
            else:
                cell.style.fill.fill_type = Fill.FILL_SOLID
                cell.style.fill.start_color.index = color

    DEFAULT_LABEL_STYLE = {
        'color': Color.WHITE,
        'bold': True,
        'halign': Alignment.HORIZONTAL_RIGHT
    }

    DEFAULT_CELL_STYLE = {
        'formula': True,
        'bold': True,
        'color': Color.WHITE,
        'format': NumberFormat.FORMAT_CURRENCY_USD_SIMPLE
    }

    def _make_totals(self, start_row, label = None, label_pos = None, cell_style = None, **label_opts):
        if cell_style is not None:
            cell_style = dict(self.DEFAULT_CELL_STYLE.items() + cell_style.items())
        else:
            cell_style = self.DEFAULT_CELL_STYLE

        for idx in xrange(1, len(self.TOTALS) + 1):
           col = self.COLUMN_UNITPRICE + idx

           self.cell(
               col, self.row,
               '=SUBTOTAL(9,%s%d:%s%d)' % (
                   gcl(col), start_row,
                   gcl(col), self.row - 1
               ),
               **cell_style
           )

        if label is not None:
            self.cell(self.COLUMN_UNITPRICE if label_pos is None else label_pos, self.row, value=label, **dict(self.DEFAULT_LABEL_STYLE.items() + label_opts.items()))

    def section_end(self):
        self._row_color()
        self._make_totals(self.section_row, label='%s TOTAL' % (self.location_name,))
        self._row_color(row=self.row + 1)
        self.row += 2
        self.ws.page_breaks.append(self.row - 1)

    def _make_borders(self):
        for i in xrange(1, self.row):
            self._row_borders(i)

    def _set_widths(self):
        for i, w in enumerate(self.COLUMN_WIDTH):
            c = chr(ord('A') + i)

            if c not in self.ws.column_dimensions:
                break

            self.ws.column_dimensions[c].width = w

    def save(self, fname):
        self._set_widths()
        self._make_borders()

        if fname is None:
            return save_virtual_workbook(self.wb)
        else:
            return self.wb.save(fname)

    def _row_borders(self, row = None, border = Border.BORDER_THIN):
        if row is None:
            row = self.row

        for i in xrange(1, len(self.HEADER) + 1):
            self.cell(i, row, border=border)

    def grand_total(self, label = 'TOTAL FOR ALL LOCATIONS', **totals_opts):
        pass
        # not needed anymore
        #self._row_color(None, self.row - 1)
        #self.row += 1
        #self._add_header(self.COLUMN_UNITPRICE, self.TOTALS)
        #self.row += 1
        #self._make_totals(1, label=label, **totals_opts)
        #self._row_color()
        #self.row += 1

class OrdersReport(ReportFormatter):
    COLUMN_ITEM      = 1
    COLUMN_PARTNO    = 2
    COLUMN_DESCR     = 3
    COLUMN_QUANTITY  = 4
    COLUMN_UNITPRICE = 5
    COLUMN_EXTENDED  = 6
    COLUMN_SHIPPING  = 7
    COLUMN_HST       = 8
    COLUMN_TOTAL     = 9

    HEADER = ['Item', 'Part number', 'Description', 'Qty', 'Unit Price', 'Extended', 'Shipping', 'HST', 'Total']
    COLUMN_WIDTH = [53, 18, 11, 10, 7, 11, 10, 11]

    LIGHT_BORDER = 'BFBFBF'

    def __init__(self):
        ReportFormatter.__init__(self)

    def section(self, location):
        self._row_color(self.LIGHT_BORDER)
        self.cell(self.COLUMN_ITEM, self.row, value=location, bold=True)

        self.row += 1

        self._remember_section(location)

    def order(self, order):
        self.cell(self.COLUMN_ITEM, self.row, value='Order')
        self.cell(self.COLUMN_ITEM, self.row + 1, value='Date')

        self.cell(self.COLUMN_PARTNO, self.row, value=order.name, bold=True)
        self.cell(self.COLUMN_PARTNO, self.row + 1, value=order.date, format=NumberFormat.FORMAT_DATE_XLSX14)

        self.row += 3
        self._row_borders()
        self._add_header()
        self.row += 1

    def items(self, items):
        ReportFormatter.items(self, items)
        self.row += 1

    def item(self, item):
        self.cell(self.COLUMN_DESCR, self.row).value = item.inventory_history.product.description
        self._row_borders()
        ReportFormatter.item(self, item)
        self.cell(self.COLUMN_ITEM, self.row - 1, wrap_text=False)

    TOTAL_CELL_STYLE = {
        'color': Color.BLACK
    }

    def section_end(self):
        self.row -= 1
        self._row_borders()
        self._row_color(self.LIGHT_BORDER)
        self._make_totals(
            self.section_row,
            label='%s ORDER TOTAL' % (self.location_name,),
            color=Color.BLACK,
            label_pos=self.COLUMN_ITEM,
            halign=Alignment.HORIZONTAL_LEFT,
            cell_style=self.TOTAL_CELL_STYLE
        )
        self._row_color('606060', self.row + 1)
        self.row += 2

    def _make_borders(self):
        pass
    
    GRAND_TOTAL_CELL = {
        'color': Color.WHITE
    }

    def grand_total(self, label = 'TOTAL FOR ALL LOCATIONS'):
        ReportFormatter.grand_total(self, label, cell_style=self.GRAND_TOTAL_CELL, label_pos=self.COLUMN_ITEM, halign=Alignment.HORIZONTAL_LEFT)
        self._row_borders(self.row - 2)
        self._row_borders(self.row - 1)
